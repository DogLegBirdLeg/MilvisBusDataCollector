from openpyxl import load_workbook
from datetime import time

load_wb = load_workbook('C:/projects/collector/부산대 시간표.xlsx', data_only=True)

load_ws = load_wb['대학공휴일']

start_column = 11
station_depart_time_column = 14
line_id_column = 17


class Data:
    def __init__(self, start, station_depart_time: time, line_id: str):
        self.start = start
        self.station_depart_time = station_depart_time.strftime('%H:%M')
        self.line_id = line_id.replace('번', '')


datas = []
for i in range(6, 21+1):
    start_cell = load_ws.cell(row=i, column=start_column)
    station_depart_time_cell = load_ws.cell(row=i, column=station_depart_time_column)
    line_id_cell = load_ws.cell(row=i, column=line_id_column)

    data = Data(start_cell.value, station_depart_time_cell.value, line_id_cell.value)

    datas.append(data)


for data in datas:
    print(f"'{data.station_depart_time}': '{data.start}/{data.line_id}',")
